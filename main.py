from fastapi import FastAPI, Request, Header, Depends, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from pydantic import BaseModel
from collections import Counter
from datetime import datetime, timedelta
import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import requests
import json
import os
import io
import threading
import time
from fastapi import Body
import platform
import openpyxl
from fastapi import Depends
import csv
from io import BytesIO
from fastapi import APIRouter
from typing import List
from typing import List, Dict
from datetime import datetime, timedelta, timezone
from datetime import datetime, time as dtime
from typing import List, Union
from datetime import datetime, time as dtime
from typing import Optional
import pytz

SUPERVISOR_FILE = "supervisors.json"
LINEWALKER_FILE = "linewalkers.json"
NPV_FILE = "npv.json"
RESET_DURATION_HOURS = 9

app = FastAPI()
sent_count = 0

API_KEY = "Yj@mb51"
DB_FILE = "log.sqlite"
SETTINGS_FILE = "settings.json"
GROUP_MAP_FILE = "group_mapping.json"
EXCEL_EXPORT_FILE = "PIDS_Log_Export.xlsx"  # or a full path

settings = {}
received_messages = []
last_update_id = None

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def verify_api_key(x_api_key: str = Header(...)):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=403, detail="Unauthorized")

SETTINGS_FILE = "settings.json"
GROUP_MAP_FILE = "group_mapping.json"

import os
import json
from typing import List, Union
from fastapi import FastAPI, Depends
from fastapi.responses import JSONResponse
from pydantic import BaseModel

app = FastAPI()

# === Constants ===
# === Constants ===
BASE_DIR = os.path.dirname(__file__)
SETTINGS_FILE = os.path.join(BASE_DIR, "settings.json")
GROUP_MAP_FILE = os.path.join(BASE_DIR, "group_mapping.json")

DEFAULT_SETTINGS = {
    "BOT_TOKEN": "",
    "ALERT_SOUND": True,
    "TIMEOUT": 10
}

# Global in-memory variables
settings = {}
group_mappings = []

# === Load Settings ===
def load_settings():
    global settings
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE) as f:
                loaded = json.load(f)
        except json.JSONDecodeError:
            print("⚠️ Corrupt settings file. Resetting to default.")
            loaded = {}
        settings = {**DEFAULT_SETTINGS, **loaded}
    else:
        print("⚠️ Settings file not found. Creating with defaults.")
        settings = DEFAULT_SETTINGS.copy()

    try:
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(settings, f, indent=2)
    except Exception as e:
        print(f"❌ Could not write settings file: {e}")

    print(f"✅ Loaded settings: {settings}")
    return settings

# === Load Group Mappings ===
def load_group_mappings():
    global group_mappings
    if os.path.exists(GROUP_MAP_FILE):
        try:
            with open(GROUP_MAP_FILE) as f:
                group_mappings = json.load(f)
        except Exception as e:
            print(f"⚠️ Failed to load group mappings: {e}")
            group_mappings = []
    else:
        group_mappings = []
    print(f"✅ Loaded group mappings: {len(group_mappings)} entries")
    return group_mappings

# === Dummy Auth ===
def verify_api_key():
    return True

# === Load configs at backend startup ===
settings = load_settings()
group_mappings = load_group_mappings()


# === Models ===
class TokenUpdateRequest(BaseModel):
    token: str

class GroupMappingItem(BaseModel):
    start_ch: float
    end_ch: float
    group_name: str
    chat_id: Union[int, str]

class GroupMappingPayload(BaseModel):
    group_mappings: List[GroupMappingItem]

# === API: Update BOT Token ===
@app.post("/update_token")
def update_token(data: TokenUpdateRequest, auth=Depends(verify_api_key)):
    global settings
    try:
        if not data.token.strip():
            return JSONResponse(
                status_code=400,
                content={"status": "error", "detail": "Token cannot be empty"}
            )

        # Update in-memory settings
        settings["BOT_TOKEN"] = data.token.strip()

        # Write to file
        try:
            with open(SETTINGS_FILE, "w") as f:
                json.dump(settings, f, indent=2)
        except Exception as e:
            print(f"❌ Failed to write settings: {e}")
            return JSONResponse(
                status_code=500,
                content={"status": "error", "detail": "Failed to write settings. Possibly read-only directory."}
            )

        # ✅ Reload in-memory settings from file to ensure consistency
        load_settings()

        print("✅ BOT_TOKEN updated and reloaded:", settings["BOT_TOKEN"])
        return {"status": "success", "message": "Bot token updated"}

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "detail": str(e)}
        )


# === API: Update Group Mappings ===
@app.post("/update_group_mappings")
def update_group_mappings(payload: GroupMappingPayload, auth=Depends(verify_api_key)):
    try:
        mappings = [item.dict() for item in payload.group_mappings]
        print("📝 Updating group mappings...")
        print("📂 GROUP_MAP_FILE:", GROUP_MAP_FILE)

        # Step 1: Save to file
        try:
            with open(GROUP_MAP_FILE, "w") as f:
                json.dump(mappings, f, indent=2)
        except Exception as e:
            print(f"❌ Failed to write group mappings: {e}")
            return JSONResponse(
                status_code=500,
                content={"status": "error", "detail": "Write failed. Possibly read-only file system."}
            )

        # ✅ Step 2: Re-load from file to confirm and sync in memory
        reloaded = load_group_mappings()
        print("✅ Group mappings reloaded:", reloaded)

        return {
            "status": "success",
            "message": "Group mappings updated",
            "updated": reloaded
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "detail": str(e)}
        )




def log_message_sqlite(data):
    now = datetime.now()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        INSERT INTO sent_logs (date, time, od, ch, section, linewalker, supervisor, npv)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        data.get("OD", ""),
        data.get("CH", ""),
        data.get("Section", ""),
        data.get("LineWalker", ""),
        data.get("Supervisor", ""),
        data.get("NPV", "")
    ))
    conn.commit()
    conn.close()

@app.get("/receive")
def get_received_logs(limit: int = 100):
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT timestamp, linewalker, message, user 
            FROM received_messages 
            ORDER BY timestamp DESC 
            LIMIT ?
        """, (limit,))
        rows = c.fetchall()
        conn.close()

        return [
            {
                "User": row[3],  # user
                "Message": row[2],  # message
                "Time": row[0].split(" ")[1] if " " in row[0] else row[0]  # extract HH:MM:SS
            }
            for row in rows
        ]

    except Exception as e:
        print(f"[DB Error] {e}")
        return {"error": str(e)}
# ========== Settings and Linewalkers ==========

def load_linewalkers():
    if os.path.exists(LINEWALKER_FILE):
        with open(LINEWALKER_FILE) as f:
            return json.load(f)
    return []

def save_linewalkers(data):
    with open(LINEWALKER_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def refresh_linewalkers():
    global linewalker_data
    linewalker_data = load_linewalkers()

linewalker_data = load_linewalkers()

# ========== Section Data ==========
def load_section_data():
    section_data = {}
    try:
        df_master = pd.read_csv("OD_CH_Master.csv")
        df_master = df_master.dropna(subset=["Section", "OD", "CH", "Diff"])

        for section in df_master["Section"].unique():
            df_section = df_master[df_master["Section"] == section].copy()
            df_section = df_section.sort_values("OD").reset_index(drop=True)
            section_data[section] = df_section

    except Exception as e:
        print(f"❌ Error loading OD_CH_Master.csv: {e}")

    return section_data

# ========== Interpolation ==========
def interpolate_ch(df, od):
    ch_matches = []
    for i in range(len(df) - 1):
        od1 = df.loc[i, "OD"]
        od2 = df.loc[i + 1, "OD"]
        ch1 = df.loc[i, "CH"]
        ch2 = df.loc[i + 1, "CH"]
        diff = df.loc[i, "Diff"]
        if od1 <= od <= od2 and diff != 0:
            od_diff = od - od1
            ch = ch1 + ((ch2 - ch1) * od_diff / diff)
            ch_matches.append(round(ch, 3))
    return ch_matches  # Always returns a list

def interpolate_od(df, ch):
    for i in range(len(df) - 1):
        ch1 = df.loc[i, "CH"]
        ch2 = df.loc[i + 1, "CH"]
        od1 = df.loc[i, "OD"]
        od2 = df.loc[i + 1, "OD"]
        if ch1 <= ch <= ch2:
            interpolated = od1 + ((ch - ch1) * (od2 - od1)) / (ch2 - ch1)
            return round(interpolated)
    return None

def get_linewalker_by_ch(ch):
    for entry in linewalker_data:
        if entry["start_ch"] <= ch <= entry["end_ch"]:
            return entry["line_walker"]
    return None


# ========== Main API ==========
@app.get("/calculate_ch_for_section")
def calculate_ch_for_section(section: str, od: float):
    print(f"[CH Lookup] Section={section}, OD={od}")

    section_data = load_section_data()
    df = section_data.get(section)
    if df is None:
        return {"error": f"Section '{section}' not found."}

    ch_matches = interpolate_ch(df, od)

    if not ch_matches:
        return {"error": "OD out of range or no valid interpolation found."}

    if len(ch_matches) > 1:
        print(f"[Multiple CHs] Found: {ch_matches}")
        return ch_matches

    ch_val = ch_matches[0]
    lw = get_linewalker_by_ch(ch_val)
    if not lw:
        return {"error": "Line walker not found for CH."}

    return {
        "ch": ch_val,
        "line_walker": lw
    }



@app.get("/convert/ch-to-od")
def convert_ch_to_od(section: str, ch: float):
    section_data = load_section_data()
    df = section_data.get(section)
    if df is None:
        return {"error": f"Section '{section}' not found."}
    od = interpolate_od(df, ch)
    if od is None:
        return {"error": "CH out of range."}
    return {"od": od}

@app.get("/get_sections")
def get_sections():
    try:
        section_data = load_section_data()
        return list(section_data.keys())
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})



class AlertPayload(BaseModel):
    od: float
    ch: float
    section: str
    line_walker: str
    supervisor: Optional[str] = ""
    npv: Optional[str] = ""


def load_group_mappings():
    if os.path.exists(GROUP_MAP_FILE):
        with open(GROUP_MAP_FILE) as f:
            return json.load(f)
    return []

def is_in_time_range(start: dtime, end: dtime, now: dtime):
    if start < end:
        return start <= now <= end
    else:
        return now >= start or now <= end  # For overnight ranges

def get_chat_id_by_ch_and_time(ch):
    group_mappings = load_group_mappings()
    now_ist = datetime.now(pytz.timezone("Asia/Kolkata"))
    now = now_ist.time()  # Ensure IST timezone here too

    for entry in group_mappings:
        try:
            start_ch = float(entry["start_ch"])
            end_ch = float(entry["end_ch"])
            window = entry.get("time_window", "day").lower()

            if start_ch <= ch <= end_ch:
                if window == "day" and is_in_time_range(dtime(6, 30), dtime(19, 0), now):
                    return str(entry["chat_id"])
                elif window == "night" and is_in_time_range(dtime(19, 45), dtime(5, 45), now):
                    return str(entry["chat_id"])
        except Exception as e:
            print(f"⚠️ Error in group mapping parsing: {entry} → {e}")
    return None


@app.post("/send_alert")
def send_alert(payload: AlertPayload):
    # Ensure IST timezone
    now_ist = datetime.now(pytz.timezone("Asia/Kolkata"))
    current_time = now_ist.time()

    print("📥 Received payload:", payload.dict())
    print("⏰ Current IST time:", current_time)

    line_walker = payload.line_walker
    section = payload.section
    ch = payload.ch
    od = payload.od
    supervisor = payload.supervisor or ""
    npv = payload.npv or ""

    # Time windows
    morning_start = dtime(6, 0)
    noon_end = dtime(13, 45)
    afternoon_start = dtime(13, 45)
    afternoon_end = dtime(16, 0)
    evening_start = dtime(16, 0)
    evening_end = dtime(18, 45)
    night_start = dtime(18, 45)
    night_end = dtime(6, 0)

    # 🧭 Fallback Supervisor
    if not supervisor:
        for s in load_supervisors():
            if s.get("section") == section:
                supervisor = s.get("supervisor", "")
                break
    print(f"👨‍💼 Supervisor for {section}:", supervisor)

    # 🧭 Fallback NPV
    if not npv:
        for n in load_npv():
            if n.get("section") == section:
                npv = n.get("npv", "")
                break
    print(f"🛡️ NPV for {section}:", npv)

    # Determine sender
    if is_in_time_range(morning_start, noon_end, current_time):
        print("🌅 Morning window matched")
        sender_info = f"🚶‍➡️लाइन वॉकर: {line_walker}"
    elif is_in_time_range(afternoon_start, afternoon_end, current_time):
        print("🌞 Afternoon window matched")
        sender_info = f"🚶‍➡️लाइन वॉकर: {line_walker}\n👨‍💼Sup: {supervisor}"
    elif is_in_time_range(evening_start, evening_end, current_time):
        print("🌇 Evening window matched")
        sender_info = f"👨‍💼Sup: {supervisor}"
    elif is_in_time_range(night_start, night_end, current_time):
        print("🌃 Night window matched")
        sender_info = f"🛡️NPV: {npv}"
    else:
        print("🚫 No matching time window")
        sender_info = "🚨Contact unavailable for this time window"

    # Final message
    msg = (
        "🔔 अलार्म सूचना 🔔\n"
        f"⏱️समय: {now_ist.strftime('%d-%m-%Y %H:%M:%S')}\n"
        f"🔎OD: {od}\n"
        f"📍CH: {ch}\n"
        f"📈सेक्शन: {section}\n"
        f"{sender_info}"
    )
    print("📨 Final message:\n", msg)

    # Get Telegram chat ID
    chat_id = get_chat_id_by_ch_and_time(ch)
    print("📬 Selected chat_id:", chat_id)

    if not chat_id:
        return JSONResponse(
            status_code=404,
            content={"status": "error", "detail": f"No group available for CH {ch} at this time"}
        )

    bot_token = settings.get("BOT_TOKEN", "")
    print("🔐 Using BOT_TOKEN:", bot_token[:10] + "..." if bot_token else "❌ Missing")

    if not bot_token:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "detail": "BOT_TOKEN is missing from settings"}
        )

    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"

    try:
        res = requests.post(url, json={"chat_id": chat_id, "text": msg})
        print("📡 Telegram response:", res.status_code, res.text)

        if res.status_code == 200:
            log_message_sqlite({
                "OD": od,
                "CH": ch,
                "Section": section,
                "LineWalker": line_walker,
                "Supervisor": supervisor,
                "NPV": npv
            })
            return JSONResponse(
                status_code=200,
                content={
                    "status": "success",
                    "message_id": res.json().get("result", {}).get("message_id", None)
                }
            )
        else:
            return JSONResponse(
                status_code=500,
                content={"status": "error", "detail": f"Telegram API returned {res.status_code}: {res.text}"}
            )

    except Exception as e:
        print("❌ Exception while sending alert:", str(e))
        return JSONResponse(
            status_code=500,
            content={"status": "error", "detail": f"Exception while sending alert: {str(e)}"}
        )


def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Table: Sent Logs – now includes supervisor and npv
    c.execute('''
        CREATE TABLE IF NOT EXISTS sent_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            time TEXT,
            od REAL,
            ch TEXT,
            section TEXT,
            linewalker TEXT,
            supervisor TEXT,
            npv TEXT
        )
    ''')

    # Table: Received Messages
    c.execute('''
        CREATE TABLE IF NOT EXISTS received_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            linewalker TEXT,
            message TEXT,
            user TEXT
        )
    ''')

    # Table: Duty Status
    c.execute('''
        CREATE TABLE IF NOT EXISTS duty_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            linewalker TEXT,
            duty_on TEXT,
            duty_off TEXT
        )
    ''')

    conn.commit()
    conn.close()


init_db()


@app.post("/webhook")
async def webhook(request: Request, background_tasks: BackgroundTasks):
    data = await request.json()
    background_tasks.add_task(handle_webhook, data)
    return {"status": "received"}

def handle_webhook(data):
    try:
        msg = data.get("message", {})
        text = msg.get("text", "").strip()
        user = msg.get("from", {}).get("first_name", "Unknown")

        if text:
            log_duty_status_from_message(user, text, user)

    except Exception as e:
        print("Webhook error:", e)



def log_duty_status_from_message(linewalker, message, user):
    msg_lower = message.lower()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    duty_on_msg = message if "on" in msg_lower and "off" not in msg_lower else None
    duty_off_msg = message if "off" in msg_lower else None

    if duty_on_msg or duty_off_msg:
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute('''
                INSERT INTO duty_status (timestamp, linewalker, duty_on, duty_off)
                VALUES (?, ?, ?, ?)
            ''', (timestamp, linewalker, duty_on_msg, duty_off_msg))
            conn.commit()
            print(f"[LOG] Duty status logged for {linewalker}")
        except Exception as e:
            print(f"[Log Error] Failed to log duty status: {e}")
        finally:
            conn.close()
    else:
        log_received_message(linewalker, message, user)


def log_received_message(linewalker, message, user):
    now = datetime.now()
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''
            INSERT INTO received_messages (timestamp, linewalker, message, user)
            VALUES (?, ?, ?, ?)
        ''', (
            now.strftime("%Y-%m-%d %H:%M:%S"),
            linewalker,
            message,
            user
        ))
        conn.commit()
        print(f"[✓] Logged general message from {linewalker}")
    except Exception as e:
        print(f"[Log Error] Failed to log received message: {e}")
    finally:
        conn.close()


def clear_duty_status_if_due():
    last_cleared_date = None
    while True:
        now = datetime.now()
        if now.strftime("%H:%M") == "06:30" and last_cleared_date != now.strftime("%Y-%m-%d"):
            try:
                conn = sqlite3.connect(DB_FILE)
                c = conn.cursor()
                c.execute("DELETE FROM duty_status")
                conn.commit()
                conn.close()
                print(f"[✓] Duty_Status auto-cleared at 06:30 on {now.strftime('%Y-%m-%d')}")
                last_cleared_date = now.strftime("%Y-%m-%d")
            except Exception as e:
                print(f"[!] Error clearing Duty_Status: {e}")
        time.sleep(60)


threading.Thread(target=clear_duty_status_if_due, daemon=True).start()

@app.get("/set_webhook")
def set_webhook():
    try:
        url = f"https://api.telegram.org/bot{settings['BOT_TOKEN']}/setWebhook"
        webhook_url = "https://pids-backend.onrender.com"
        res = requests.get(url, params={"url": webhook_url})
        return res.json()
    except Exception as e:
        return {"status": "error", "detail": str(e)}

# ============== View Logs  ============

@app.get("/get_logs")
def get_logs():
    try:
        conn = sqlite3.connect(DB_FILE)
        df_recv = pd.read_sql("SELECT * FROM received_messages", conn)
        df_duty = pd.read_sql("SELECT * FROM duty_status", conn)
        df_sent = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        return {
            "received": df_recv.to_dict(orient="records"),
            "duty": df_duty.to_dict(orient="records"),
            "sent": df_sent.to_dict(orient="records")
        }

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# ============== Download Logs  ============

@app.get("/download_logs")
def download_logs():
    try:
        # ✅ Step 1: Connect and read data
        conn = sqlite3.connect(DB_FILE)
        df_recv = pd.read_sql("SELECT * FROM received_messages", conn)
        df_duty = pd.read_sql("SELECT * FROM duty_status", conn)
        df_sent = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        # ✅ Step 2: Check if all DataFrames are empty
        if df_recv.empty and df_duty.empty and df_sent.empty:
            raise HTTPException(status_code=404, detail="No logs available to export.")

        # ✅ Step 3: Create Excel workbook and add sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Received Logs"

        if not df_recv.empty:
            ws1.append(df_recv.columns.tolist())
            for row in df_recv.itertuples(index=False):
                ws1.append(list(row))
        else:
            ws1.append(["No received logs available"])

        # Sheet 2: Duty Status
        ws2 = wb.create_sheet(title="Duty Status")
        if not df_duty.empty:
            ws2.append(df_duty.columns.tolist())
            for row in df_duty.itertuples(index=False):
                ws2.append(list(row))
        else:
            ws2.append(["No duty status logs available"])

        # Sheet 3: Sent Logs
        ws3 = wb.create_sheet(title="Sent Logs")
        if not df_sent.empty:
            ws3.append(df_sent.columns.tolist())
            for row in df_sent.itertuples(index=False):
                ws3.append(list(row))
        else:
            ws3.append(["No sent logs available"])

        # ✅ Step 4: Save Excel file
        wb.save(EXCEL_EXPORT_FILE)

        # ✅ Step 5: Return file
        return FileResponse(
            path=EXCEL_EXPORT_FILE,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename="PIDS_Log_Export.xlsx"
        )

    except Exception as e:
        print("❌ Error generating Excel:", str(e))
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ================ Analytics Charts from SQLite ===============
@app.get("/analytics/scatter_chart")
def get_scatter_chart():
    try:
        conn = sqlite3.connect(DB_FILE)
        df = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        df['datetime'] = pd.to_datetime(df['date'] + ' ' + df['time'])
        now = datetime.now()
        today_630 = now.replace(hour=6, minute=30, second=0, microsecond=0)
        if now < today_630:
            today_630 -= timedelta(days=1)
        tomorrow_630 = today_630 + timedelta(days=1)
        df = df[(df['datetime'] >= today_630) & (df['datetime'] < tomorrow_630)]

        fig, ax = plt.subplots(figsize=(11, 6))
        for section, group in df.groupby("section"):
            ax.scatter(group['datetime'], group['ch'].astype(float), label=section, s=40, alpha=0.8)

        ax.set_title("📊 Chainage vs Time (Section-wise, Last 24 Hours)", fontsize=14)
        ax.set_xlabel("Time", fontsize=12)
        ax.set_ylabel("Chainage", fontsize=12)
        ax.legend(title="Section", loc="best")
        ax.grid(True)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        fig.autofmt_xdate()

        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        plt.close(fig)
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/analytics/grouping_chart")
def get_grouping_chart(by: str = "section"):
    try:
        valid_fields = {
            "section": "section",
            "linewalker": "linewalker",
    
        }
        if by not in valid_fields:
            raise HTTPException(status_code=400, detail=f"Invalid group type. Use one of: {', '.join(valid_fields.keys())}")

        conn = sqlite3.connect(DB_FILE)
        df = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        df['datetime'] = pd.to_datetime(df['date'] + ' ' + df['time'])
        now = datetime.now()
        today_630 = now.replace(hour=6, minute=30, second=0, microsecond=0)
        if now < today_630:
            today_630 -= timedelta(days=1)
        tomorrow_630 = today_630 + timedelta(days=1)
        df = df[(df['datetime'] >= today_630) & (df['datetime'] < tomorrow_630)]

        counts = df[by].value_counts()

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(counts.index.astype(str), counts.values, color='teal')
        ax.set_title(f"Alarm Count by {by.capitalize()} (Last 24 Hours)", fontsize=14)
        ax.set_ylabel("Count")
        ax.set_xlabel(by.capitalize())
        ax.set_xticks(range(len(counts)))
        ax.set_xticklabels(counts.index.astype(str), rotation=45, ha='right')
        ax.grid(axis="y")

        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format="png")
        plt.close(fig)
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== Linewalker Management ==========

class LineWalkerItem(BaseModel):
    start_ch: float
    end_ch: float
    line_walker: str

@app.get("/view_linewalkers")
def view_linewalkers():
    return load_linewalkers()

@app.post("/edit_linewalkers")
def edit_linewalkers(data: list[LineWalkerItem]):
    try:
        data_dicts = [item.dict() for item in data]
        save_linewalkers(data_dicts)

        global linewalker_data
        linewalker_data = data_dicts

        return {"status": "updated"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

def save_linewalkers(data):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for item in data:
        item["saved_at"] = now_str

    with open(LINEWALKER_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def load_linewalkers():
    if not os.path.exists(LINEWALKER_FILE):
        return []
    with open(LINEWALKER_FILE) as f:
        return json.load(f)

@app.get("/refresh_linewalkers")
def refresh_linewalkers_api():
    refreshed = load_linewalkers()
    return {"status": "refreshed", "count": len(refreshed)}


class SupervisorItem(BaseModel):
    start_ch: float
    end_ch: float
    supervisor: str

@app.get("/view_supervisors")
def view_supervisors():
    return load_supervisors()

@app.post("/edit_supervisors")
def edit_supervisors(data: list[SupervisorItem] = Body(...)):
    try:
        data_dicts = [item.dict() for item in data]
        save_supervisors(data_dicts)
        return {"status": "updated"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

def save_supervisors(data):
    with open(SUPERVISOR_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def load_supervisors():
    if not os.path.exists(SUPERVISOR_FILE):
        return []
    with open(SUPERVISOR_FILE) as f:
        return json.load(f)



class NPVItem(BaseModel):
    start_ch: float
    end_ch: float
    npv: str

@app.get("/view_npv")
def view_npv():
    return load_npv()

@app.post("/edit_npv")
def edit_npv(data: list[NPVItem] = Body(...)):
    try:
        data_dicts = [item.dict() for item in data]
        save_npv(data_dicts)
        return {"status": "updated"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

def save_npv(data):
    with open(NPV_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def load_npv():
    if not os.path.exists(NPV_FILE):
        return []
    with open(NPV_FILE) as f:
        return json.load(f)


def save_permanent_contact(role: str, ch: float, name: str):
    if not name or not ch:
        return

    if role == "linewalker":
        filepath = LINEWALKER_FILE
        key = "line_walker"
    elif role == "supervisor":
        filepath = SUPERVISOR_FILE
        key = "supervisor"
    elif role == "npv":
        filepath = NPV_FILE
        key = "npv"
    else:
        return

    if not os.path.exists(filepath):
        data = []
    else:
        with open(filepath) as f:
            data = json.load(f)

    # Try to update existing CH range
    updated = False
    for entry in data:
        if entry["start_ch"] <= ch <= entry["end_ch"]:
            entry[key] = name
            updated = True
            break

    if not updated:
        # Append new CH range if not matching
        data.append({
            "start_ch": ch,
            "end_ch": ch + 0.001,  # minimal range
            key: name
        })

    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)

@app.post("/reset_all_roles")
def reset_all_roles():
    result = {}

    # === Reset Linewalkers ===
    try:
        if os.path.exists(LINEWALKER_FILE):
            with open(LINEWALKER_FILE, "r") as f:
                lw_data = json.load(f)

            for item in lw_data:
                item["line_walker"] = "-"
                item["saved_at"] = None

            with open(LINEWALKER_FILE, "w") as f:
                json.dump(lw_data, f, indent=2)

            global linewalker_data
            linewalker_data = lw_data

            result["linewalkers"] = len(lw_data)
        else:
            result["linewalkers"] = "file_not_found"
    except Exception as e:
        result["linewalkers"] = f"error: {str(e)}"

    # === Reset Supervisors ===
    try:
        if os.path.exists(SUPERVISOR_FILE):
            with open(SUPERVISOR_FILE, "r") as f:
                sv_data = json.load(f)

            for item in sv_data:
                item["supervisor"] = "-"

            with open(SUPERVISOR_FILE, "w") as f:
                json.dump(sv_data, f, indent=2)

            # Optional: global supervisor_data
            result["supervisors"] = len(sv_data)
        else:
            result["supervisors"] = "file_not_found"
    except Exception as e:
        result["supervisors"] = f"error: {str(e)}"

    # === Reset NPVs ===
    try:
        if os.path.exists(NPV_FILE):
            with open(NPV_FILE, "r") as f:
                npv_data = json.load(f)

            for item in npv_data:
                item["npv"] = "-"

            with open(NPV_FILE, "w") as f:
                json.dump(npv_data, f, indent=2)

            # Optional: global npv_data
            result["npvs"] = len(npv_data)
        else:
            result["npvs"] = "file_not_found"
    except Exception as e:
        result["npvs"] = f"error: {str(e)}"

    return {
        "status": "reset",
        "details": result
    }



@app.get("/ping")
def ping():
    return {"status": "ok"}
@app.get("/")
def root():
    return {"message": "✅ PIDS Alert Backend is Running"}

